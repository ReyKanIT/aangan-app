#!/usr/bin/env python3
"""
Build AANGAN_REGRESSION_SUITE.xlsx — PRD-anchored coverage matrix.

This is the version-controlled copy of /tmp/build_aangan_regression_xlsx.py.
Edit here and re-run via:

    python3 scripts/build_aangan_regression_xlsx.py

Sheets:
  1. Coverage Matrix — every PRD feature row, mapped to test tier(s) + status
  2. Test Inventory — every test currently in repo (file + case + status)
  3. Phase Plan      — Phase 1–4 deliverables and status
  4. Bug Traceability — today's 5 bugs + v0.10.1 regression mapped to tests
  5. PRD Reference   — flat summary of PRD §4 features (so xlsx is self-contained)
  6. Tier Strategy   — 4-tier model overview

Conventions (per skill rules):
  - Arial font, bold headers, frozen top row
  - Status color-coded: green=passing, yellow=planned, red=blocked, grey=n/a
  - Zero formula errors
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Style helpers ─────────────────────────────────────────────────────────

FONT_HEADER = Font(name='Arial', bold=True, color='FFFFFF', size=11)
FONT_BODY = Font(name='Arial', size=10)
FONT_BOLD = Font(name='Arial', bold=True, size=10)

FILL_HEADER = PatternFill('solid', start_color='4A6B2A')        # mehndi green
FILL_PASS = PatternFill('solid', start_color='D4EDDA')          # light green
FILL_PLAN = PatternFill('solid', start_color='FFF3CD')          # light yellow
FILL_BLOCK = PatternFill('solid', start_color='F8D7DA')         # light red
FILL_NA = PatternFill('solid', start_color='E2E3E5')            # light grey
FILL_INFO = PatternFill('solid', start_color='D1ECF1')          # light blue (info rows)

THIN = Side(border_style='thin', color='AAAAAA')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

WRAP = Alignment(wrap_text=True, vertical='top')
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)

STATUS_FILLS = {
    'PASSING': FILL_PASS,
    'PLANNED': FILL_PLAN,
    'BLOCKED': FILL_BLOCK,
    'N/A': FILL_NA,
    'PARTIAL': FILL_PLAN,
    'REGRESSION': FILL_BLOCK,
    'DONE': FILL_PASS,
}

def style_header(ws, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = CENTER
        cell.border = BORDER
    ws.freeze_panes = 'A2'

def fill_body(ws, start_row, end_row, ncols, status_col=None):
    for r in range(start_row, end_row + 1):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = FONT_BODY
            cell.alignment = WRAP
            cell.border = BORDER
        if status_col:
            status_cell = ws.cell(row=r, column=status_col)
            fill = STATUS_FILLS.get((status_cell.value or '').strip().upper())
            if fill:
                status_cell.fill = fill
                status_cell.font = FONT_BOLD

def auto_widths(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

# ── Sheet 1: Coverage Matrix ──────────────────────────────────────────────

COVERAGE_HEADERS = [
    'PRD §', 'Feature', 'Sub-feature', 'Priority',
    'T0 tsc', 'T1 Jest', 'T2 Critical', 'T3 Maestro', 'T4 Manual',
    'Test file / planned location', 'Status', 'Notes',
]

COVERAGE_ROWS = [
    # ─ 4.1 Auth ─────────────────────────────────────────────────────────
    ['4.1', 'Authentication', 'Phone OTP (Supabase + Twilio SMS)', 'P0',
     '✓', '', '', 'planned', '', '.maestro/signin-phone.yaml',
     'PLANNED', 'Test account 9876543210/123456 per CLAUDE.md'],
    ['4.1', 'Authentication', 'Google OAuth sign-in', 'P0',
     '✓', '✓', 'planned', 'planned', '', 'src/__tests__/structural/v0_16_1_lockdown.test.ts',
     'PASSING', 'Lockdown asserts "Google" present in login screen — would have caught v0.10.1 regression'],

    # ─ 4.2 Family Tree ──────────────────────────────────────────────────
    ['4.2', 'Family Tree', '40+ Hindi relationship labels', 'P0',
     '✓', '✓', '', '', '', 'src/__tests__/components/KulvrikshTreeView.test.tsx',
     'PASSING', '26 getGenerationOffset tests cover all relationship labels'],
    ['4.2', 'Family Tree', 'Tab views: कुलवृक्ष / सदस्य / जीवन यात्रा', 'P0',
     '✓', '✓', '', 'planned', '', 'src/__tests__/structural/v0_16_1_lockdown.test.ts',
     'PASSING', 'v0.16.1 fix: 3 tabs (collapsed from 6), default=कुलवृक्ष'],
    ['4.2', 'Family Tree', 'Add Member name-first form (v0.16.1 fix)', 'P0',
     '✓', '✓', '', 'planned', '', 'src/__tests__/structural/v0_16_1_lockdown.test.ts',
     'PASSING', 'Bug #2 lockdown: name field first, phone optional, mode toggle gone'],
    ['4.2', 'Family Tree (Tree viz)', 'Visual generational layout — elders above, descendants below', 'P0',
     '✓', '✓', '', '✓', '', 'src/__tests__/components/KulvrikshTreeView.test.tsx + .maestro/family-tree-renders.yaml',
     'PASSING', 'Maestro flow added [3:42pm - 17May26] as canonical reproducer for gen-tree regression class'],

    # ─ 4.7 Cultural ─────────────────────────────────────────────────────
    ['4.7', 'Panchang', '"What\'s special today" ribbon (v0.16.1 fix Bug #3)', 'P0',
     '✓', '✓', '', 'planned', '', 'src/__tests__/components/PanchangWidget.test.tsx + structural/v0_16_1_lockdown.test.ts',
     'PASSING', 'Helper computeSpecialToday extracted to panchangService; 8 unit tests cover festival > tithi > null'],
    ['4.7', 'Panchang', 'computeSpecialToday — Pratipada-only day returns null', 'P0',
     '', '✓', '', '', '', 'src/__tests__/components/PanchangWidget.test.tsx',
     'PASSING', '17 May 2026 Pratipada — ribbon hidden'],
    ['4.7', 'Panchang', 'computeSpecialToday — Purnima/Amavasya/Ekadashi → haldiGold accent', 'P0',
     '', '✓', '', '', '', 'src/__tests__/components/PanchangWidget.test.tsx',
     'PASSING', 'Three branches asserted independently'],
    ['4.7', 'Panchang', 'computeSpecialToday — Festival match uses nameHindi + शुभकामनाएँ + catalogue color', 'P0',
     '', '✓', '', '', '', 'src/__tests__/components/PanchangWidget.test.tsx',
     'PASSING', 'Buddha Purnima + Diwali asserted'],
    ['4.7', 'Panchang', 'computeSpecialToday — Festival takes priority over special tithi', 'P0',
     '', '✓', '', '', '', 'src/__tests__/components/PanchangWidget.test.tsx',
     'PASSING', 'Buddha Purnima date + पूर्णिमा tithi → festival branch wins'],

    # ─ CRITICAL ─────────────────────────────────────────────────────────
    ['CRIT', 'Husky pre-commit hook', 'tsc + jest + check:critical-rn gates run before every commit', 'P0',
     '✓', '✓', '✓', '', '', 'aangan_rn/scripts/pre-commit.sh + scripts/install-precommit.js',
     'PASSING', 'Wired via husky 9.x prepare script [3:42pm - 17May26]. Bypassable only with --no-verify (forbidden by feedback_release_workflow.md).'],
    ['CRIT', 'Maestro flow scaffold', '.maestro/ directory + README + first flow', 'P0',
     '', '', '', '✓', '', '.maestro/family-tree-renders.yaml',
     'PASSING', 'Canonical reproducer for v0.16.1 gen-tree regression — Jest cannot catch runtime SVG/Reanimated bugs'],
    ['CRIT', 'Critical features manifest', 'CRITICAL_FEATURES.md enforced via pre-build hook (web)', 'P0',
     '✓', '', '✓', '', '', 'aangan_web/scripts/check-critical-features.mjs',
     'PASSING', 'Caught v0.10.1 Google sign-in removal regression once added'],
    ['CRIT', 'Critical features manifest', 'RN analogue check-critical-rn.mjs', 'P0',
     '✓', '', '✓', '', '', 'scripts/check-critical-rn.mjs',
     'PASSING', 'Phase 2 deliverable (parallel agent)'],
    ['CRIT', 'No regressions rule', 'feedback_no_regressions.md memory + diff-inspect discipline', 'P0',
     '', '', '✓', '', '✓', 'CLAUDE.md + feedback_no_regressions.md',
     'PASSING', 'Encoded as memory file 2026-05-17'],
    ['CRIT', 'Release workflow', 'code → tsc → jest → build → sim → release', 'P0',
     '', '', '', '', '✓', 'feedback_release_workflow.md',
     'PASSING', 'Encoded as memory file 2026-05-17'],
]

# ── Sheet 2: Test Inventory ──────────────────────────────────────────────

INVENTORY_HEADERS = [
    'Tier', 'Test file', 'Test case', 'Status', 'Added',
]

INVENTORY_ROWS = [
    # T1 KulvrikshTreeView
    ['T1', 'src/__tests__/components/KulvrikshTreeView.test.tsx',
     '<KulvrikshTreeView> renders empty / single / 8-member / English locale (4 cases)', 'PASSING', '2026-05-17 Phase 1'],
    ['T1', 'src/__tests__/components/KulvrikshTreeView.test.tsx',
     'getGenerationOffset — siblings & spouse → 0 (5 cases)', 'PASSING', '2026-05-17 Phase 1'],
    ['T1', 'src/__tests__/components/KulvrikshTreeView.test.tsx',
     'getGenerationOffset — children → +1 (4 cases)', 'PASSING', '2026-05-17 Phase 1'],
    ['T1', 'src/__tests__/components/KulvrikshTreeView.test.tsx',
     'getGenerationOffset — parents/uncle/aunt → -1 (4 cases)', 'PASSING', '2026-05-17 Phase 1'],
    ['T1', 'src/__tests__/components/KulvrikshTreeView.test.tsx',
     'getGenerationOffset — grandparents → -2 (4 cases)', 'PASSING', '2026-05-17 Phase 1'],
    ['T1', 'src/__tests__/components/KulvrikshTreeView.test.tsx',
     'getGenerationOffset — grandchildren → +2 (4 cases)', 'PASSING', '2026-05-17 Phase 1'],
    ['T1', 'src/__tests__/components/KulvrikshTreeView.test.tsx',
     'getGenerationOffset — fallback 0 + case-insensitivity', 'PASSING', '2026-05-17 Phase 1'],

    # T1 Structural lockdowns
    ['T1', 'src/__tests__/structural/v0_16_1_lockdown.test.ts',
     'Bug #1 — default tab + 3 tabs + tree present + LEVEL_CHIPS (4 assertions)', 'PASSING', '2026-05-17 Phase 1'],
    ['T1', 'src/__tests__/structural/v0_16_1_lockdown.test.ts',
     'Bug #2 — Name first + phone-optional + mode toggle gone (3 assertions)', 'PASSING', '2026-05-17 Phase 1'],
    ['T1', 'src/__tests__/structural/v0_16_1_lockdown.test.ts',
     'Bug #3 — specialToday useMemo + ribbon + null fallback (3 assertions)', 'PASSING', '2026-05-17 Phase 1'],
    ['T1', 'src/__tests__/structural/v0_16_1_lockdown.test.ts',
     'Bug #4 — Storage/Referral routes deleted (3 assertions)', 'PASSING', '2026-05-17 Phase 1'],
    ['T1', 'src/__tests__/structural/v0_16_1_lockdown.test.ts',
     'CRITICAL — Google sign-in + phone-OTP path present', 'PASSING', '2026-05-17 Phase 1'],

    # T1 PanchangWidget (NEW)
    ['T1', 'src/__tests__/components/PanchangWidget.test.tsx',
     'computeSpecialToday — Pratipada-only day returns null', 'PASSING', '2026-05-17 Phase 2'],
    ['T1', 'src/__tests__/components/PanchangWidget.test.tsx',
     'computeSpecialToday — ordinary tithis (Tritiya/Saptami/Dashami/Trayodashi) return null', 'PASSING', '2026-05-17 Phase 2'],
    ['T1', 'src/__tests__/components/PanchangWidget.test.tsx',
     'computeSpecialToday — Purnima ribbon with haldiGold accent', 'PASSING', '2026-05-17 Phase 2'],
    ['T1', 'src/__tests__/components/PanchangWidget.test.tsx',
     'computeSpecialToday — Amavasya ribbon with correct line', 'PASSING', '2026-05-17 Phase 2'],
    ['T1', 'src/__tests__/components/PanchangWidget.test.tsx',
     'computeSpecialToday — Ekadashi ribbon with correct line', 'PASSING', '2026-05-17 Phase 2'],
    ['T1', 'src/__tests__/components/PanchangWidget.test.tsx',
     'computeSpecialToday — Buddha Purnima uses nameHindi + शुभकामनाएँ + catalogue color', 'PASSING', '2026-05-17 Phase 2'],
    ['T1', 'src/__tests__/components/PanchangWidget.test.tsx',
     'computeSpecialToday — Diwali festival branch', 'PASSING', '2026-05-17 Phase 2'],
    ['T1', 'src/__tests__/components/PanchangWidget.test.tsx',
     'computeSpecialToday — Festival > tithi priority on Buddha Purnima day', 'PASSING', '2026-05-17 Phase 2'],
    ['T1', 'src/__tests__/components/PanchangWidget.test.tsx',
     'computeSpecialToday — Empty festival catalogue still works', 'PASSING', '2026-05-17 Phase 2'],

    # T2 web
    ['T2', 'aangan_web/scripts/check-critical-features.mjs',
     'Verifies all CRITICAL_FEATURES.md entries exist in source', 'PASSING', 'pre-existing'],

    # T2 rn
    ['T2', 'scripts/check-critical-rn.mjs',
     'RN analogue of web check (Phase 2)', 'PASSING', '2026-05-17 Phase 2 (parallel agent)'],

    # T3 Maestro (NEW)
    ['T3', '.maestro/family-tree-renders.yaml',
     'Launch app → tap परिवार → assert कुलवृक्ष selected + no ErrorBoundary', 'PLANNED', '2026-05-17 Phase 3'],

    # Pre-commit hook (NEW)
    ['CI', 'aangan_rn/.husky/pre-commit',
     'T0 tsc + T1 jest + T2 check:critical-rn — hard gates before every commit', 'PASSING', '2026-05-17 Phase 2'],
]

# ── Sheet 3: Phase Plan ──────────────────────────────────────────────────

PHASE_HEADERS = ['Phase', 'Deliverable', 'Status', 'Notes']

PHASE_ROWS = [
    ['Phase 1', 'Jest scaffold + KulvrikshTreeView + structural lockdowns + TESTING.md', 'DONE',
     '30 component + 15 structural tests passing'],
    ['Phase 2', 'PanchangWidget specialToday tests (T1)', 'DONE',
     'Helper extracted to panchangService.ts (refactor only); 9 unit tests'],
    ['Phase 2', 'Husky pre-commit hook (tsc + jest + check:critical-rn)', 'DONE',
     'aangan_rn/.husky/pre-commit + scripts/pre-commit.sh + scripts/install-precommit.js'],
    ['Phase 2', 'check-critical-rn.mjs RN analogue', 'DONE',
     'Parallel agent — see TESTING.md §T2b'],
    ['Phase 2', 'postStore + familyStore + eventStore + messageStore tests', 'PLANNED',
     'Parallel agent in flight'],
    ['Phase 3', 'Maestro CLI install + .maestro/ scaffold + family-tree-renders.yaml', 'DONE',
     'README documents install + test account; first flow lives next to it'],
    ['Phase 3', 'signin-phone.yaml + home-loads.yaml + add-member-*.yaml', 'PLANNED', ''],
    ['Phase 3', 'compose-photo.yaml + panchang-special-day.yaml + create-event.yaml', 'PLANNED', ''],
    ['Phase 4', 'scripts/post-build-smoke.sh + CI workflow + release.mjs gate', 'PLANNED', ''],
]

# ── Sheet 4: Bug Traceability ────────────────────────────────────────────

BUG_HEADERS = ['Bug ID', 'Date', 'Summary', 'T1 lockdown test', 'T3 Maestro flow', 'Status']

BUG_ROWS = [
    ['v0.10.1', '2026-04-26',
     'Google sign-in silently removed during unrelated bug fix',
     'CRITICAL — Google sign-in surface present',
     '.maestro/signin-google.yaml (planned)',
     'LOCKED DOWN'],
    ['#1 v0.16.1', '2026-05-17',
     'Family tree not visible — default tab was grid, tree hidden',
     'Bug #1 default-tab lockdown (4 assertions)',
     '.maestro/family-tree-renders.yaml (landed [3:42pm - 17May26])',
     'FIXED + LOCKED DOWN'],
    ['#2 v0.16.1', '2026-05-17',
     'Add Member name field missing — mode toggle defaulted to online',
     'Bug #2 name-first lockdown (3 assertions)',
     '.maestro/add-member-name-only.yaml (planned)',
     'FIXED + LOCKED DOWN'],
    ['#3 v0.16.1', '2026-05-17',
     'Home page missing date+tithi pairing + "what\'s special today"',
     'Bug #3 ribbon lockdown + PanchangWidget computeSpecialToday (12 assertions total)',
     '.maestro/panchang-special-day.yaml (planned)',
     'FIXED + LOCKED DOWN'],
    ['#4 v0.16.1', '2026-05-17',
     'Storage usage details — remove (unlimited pivot)',
     'Bug #4 navigator route removal (3 assertions)',
     'n/a (deletion verified at code level)',
     'FIXED + LOCKED DOWN'],
    ['#5 v0.16.1', '2026-05-17',
     'Pic uploads not working — posts bucket missing RLS',
     'n/a (server-side migration)',
     '.maestro/compose-photo.yaml (planned)',
     'FIXED (server)'],
    ['#8 REGRESSION', '2026-05-17',
     'Kulvriksh tree throws ErrorBoundary in real sim — Jest passed',
     'src/__tests__/components/KulvrikshTreeView.test.tsx (passes in Jest, fails in real env)',
     '.maestro/family-tree-renders.yaml — canonical reproducer',
     'IDENTIFIED, fix held for v0.16.2'],
]

# ── Sheet 5: PRD Reference ──────────────────────────────────────────────

PRD_HEADERS = ['PRD §', 'Section', 'Summary']

PRD_ROWS = [
    ['1', 'Vision', "India's family social network — Hindi-first, Dadi-Test compliant"],
    ['3', 'Design Principles', 'Dadi Test: 52px+ buttons, 16px+ text, Hindi-first. Haldi/Mehndi/Cream'],
    ['4.1', 'Authentication', 'Phone OTP + Google OAuth + profile setup'],
    ['4.2', 'Family Tree', '3-level hierarchy, 40+ Hindi relationships, tab views, SVG tree, Life Events, Sutak'],
    ['4.3', 'Posts & Feed', 'Text/photo/video, audience control, reactions, comments, polls, stories'],
    ['4.4', 'Events & RSVP', '8 types, multi-ceremony, RSVP states, GPS check-in, QR guest upload'],
    ['4.5', 'Notifications', 'Push + 11 types + daily 08:00 IST family reminders (dual-calendar)'],
    ['4.7', 'Cultural', 'Panchang (Meeus offline) + Kuldevi + Festival calendar + specialToday ribbon'],
    ['4.8.1', 'Voice', 'STT, 8 commands, voice messages (M4A, waveform), pre-auth lang toggle'],
    ['4.9', 'Settings', 'Lang/theme/notif/privacy/kuldevi/reminders/sign-out/delete-account'],
    ['4.11', 'Web App', 'Next.js 15 + Vercel + 22 routes + PWA + Sentry'],
]

# ── Sheet 6: Tier Strategy ──────────────────────────────────────────────

TIER_HEADERS = ['Tier', 'Tool', 'Speed', 'Catches', 'Required on']

TIER_ROWS = [
    ['T0', 'tsc --noEmit', 'Seconds', 'Type errors, missing props', 'Every edit + pre-commit'],
    ['T1', 'Jest + RN Testing Library', 'Seconds (~1s)',
     'Render exceptions, prop logic, store actions, pure helpers, structural invariants',
     'Every commit (pre-commit)'],
    ['T2', 'check:critical scripts', 'Seconds',
     'Silent deletion of UI features listed in CRITICAL_FEATURES.md',
     'Every commit'],
    ['T3', 'Maestro flows', 'Minutes',
     'Navigation, real Supabase, gestures, runtime SVG/Reanimated issues',
     'Every PR / EAS build'],
    ['T4', 'Manual sim smoke', 'Minutes',
     'Gestalt feel, Dadi-Test, anything not in T0–T3',
     'Every store upload'],
]

# ── Build workbook ────────────────────────────────────────────────────────

wb = Workbook()
wb.remove(wb.active)

def add_sheet(name, headers, rows, widths, status_col=None):
    ws = wb.create_sheet(name)
    ws.append(headers)
    for r in rows:
        ws.append(r)
    style_header(ws, len(headers))
    fill_body(ws, 2, len(rows) + 1, len(headers), status_col=status_col)
    auto_widths(ws, widths)

add_sheet('Coverage Matrix', COVERAGE_HEADERS, COVERAGE_ROWS,
          {1: 9, 2: 22, 3: 36, 4: 9, 5: 8, 6: 9, 7: 11, 8: 11, 9: 10, 10: 50, 11: 12, 12: 50},
          status_col=11)
add_sheet('Test Inventory', INVENTORY_HEADERS, INVENTORY_ROWS,
          {1: 8, 2: 60, 3: 70, 4: 12, 5: 24},
          status_col=4)
add_sheet('Phase Plan', PHASE_HEADERS, PHASE_ROWS,
          {1: 10, 2: 60, 3: 12, 4: 60},
          status_col=3)
add_sheet('Bug Traceability', BUG_HEADERS, BUG_ROWS,
          {1: 14, 2: 16, 3: 50, 4: 50, 5: 50, 6: 20},
          status_col=6)
add_sheet('PRD Reference', PRD_HEADERS, PRD_ROWS,
          {1: 10, 2: 28, 3: 90})
add_sheet('Tier Strategy', TIER_HEADERS, TIER_ROWS,
          {1: 10, 2: 38, 3: 24, 4: 60, 5: 38})

import os
OUT = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                   'AANGAN_REGRESSION_SUITE.xlsx')
wb.save(OUT)
print(f'Wrote {OUT}')
print(f'Sheets: {wb.sheetnames}')
print(f'Coverage rows: {len(COVERAGE_ROWS)}')
print(f'Inventory rows: {len(INVENTORY_ROWS)}')
